from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path("data/translations-24.xlsx")
OUTPUT_PATH = Path("data/wordCards.ts")
CARD_COUNT = 15

PICTURABLE_WORDS = [
    "가게",
    "가구",
    "가방",
    "가수",
    "갈비",
    "감",
    "개",
    "거실",
    "건물",
    "고기",
    "고양이",
    "공원",
    "공책",
    "공항",
    "과일",
]

IMAGE_PATHS = {
    "가게": "/cards/shop.png",
    "가구": "/cards/furniture.png",
    "가방": "/cards/bag.png",
    "가수": "/cards/singer.png",
    "갈비": "/cards/galbi.png",
    "감": "/cards/persimmon.png",
    "개": "/cards/dog.png",
    "거실": "/cards/living-room.png",
    "건물": "/cards/building.png",
    "고기": "/cards/meat.png",
    "고양이": "/cards/cat.png",
    "공원": "/cards/park.png",
    "공책": "/cards/notebook.png",
    "공항": "/cards/airport.png",
    "과일": "/cards/fruit.png",
}

LANGUAGE_CODES = {
    "영어": "en",
    "일본어": "ja",
    "중국어": "zh",
    "프랑스어": "fr",
    "스페인어": "es",
    "아랍어": "ar",
    "몽골어": "mn",
    "베트남어": "vi",
    "타이어": "th",
    "러시아어": "ru",
    "인도네시아어": "id",
    "중국어번체": "zhHant",
    "우즈베크어": "uz",
    "카자흐어": "kk",
    "키르기스어": "ky",
    "네팔어": "ne",
    "버마어": "my",
    "크메르어": "km",
    "필리핀어": "fil",
    "힌디어": "hi",
    "벵골어": "bn",
    "독일어": "de",
    "스와힐리어": "sw",
    "하우사어": "ha",
}

POS_TO_CATEGORY = {
    "명사": "noun",
    "동사": "verb",
    "형용사": "adjective",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def display_word(value: str) -> str:
    return re.sub(r"\d+$", "", value)


def korean_syllable_count(value: str) -> int:
    return len(re.findall(r"[가-힣]", value))


def first_translation(value: str) -> str:
    first = re.split(r";|,|。|，", value, maxsplit=1)[0].strip()
    return first or value


def sheet_rows(sheet: openpyxl.worksheet.worksheet.Worksheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}

    for row in rows:
      yield headers, index, row


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    cards_by_id: dict[str, dict] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        language_code = LANGUAGE_CODES.get(sheet_name, sheet_name)
        word_header = f"{sheet_name}_단어"
        definition_header = f"{sheet_name}_뜻풀이"

        for _, index, row in sheet_rows(sheet):
            card_id = clean(row[index["id"]])
            grade = clean(row[index["등급"]])
            korean = clean(row[index["어휘"]])
            part_of_speech = clean(row[index["품사"]])
            guide = clean(row[index["길잡이말"]])

            visible_word = display_word(korean)

            if grade != "1급" or part_of_speech != "명사":
                continue

            if korean_syllable_count(visible_word) > 3:
                continue

            if visible_word not in PICTURABLE_WORDS:
                continue

            card = cards_by_id.setdefault(
                card_id,
                {
                    "id": f"word-{card_id}",
                    "korean": visible_word,
                    "sourceWord": korean,
                    "level": grade,
                    "partOfSpeech": part_of_speech,
                    "category": POS_TO_CATEGORY[part_of_speech],
                    "guide": guide,
                    "imageSrc": IMAGE_PATHS.get(visible_word, ""),
                    "translations": {},
                },
            )

            translated_word = clean(row[index[word_header]]) if word_header in index else ""
            translated_definition = (
                clean(row[index[definition_header]]) if definition_header in index else ""
            )

            card["translations"][language_code] = {
                "language": sheet_name,
                "word": translated_word,
                "definition": translated_definition,
            }

    cards_by_word = {card["korean"]: card for card in cards_by_id.values()}
    selected_cards = [cards_by_word[word] for word in PICTURABLE_WORDS[:CARD_COUNT]]

    output = (
        "export type SupportedLanguageCode =\n"
        + "".join(f'  | "{code}"\n' for code in LANGUAGE_CODES.values())
        + ";\n\n"
        + "export type WordCard = {\n"
        + "  id: string;\n"
        + "  korean: string;\n"
        + "  sourceWord: string;\n"
        + "  level: string;\n"
        + "  partOfSpeech: string;\n"
        + '  category: "noun" | "verb" | "adjective";\n'
        + "  guide: string;\n"
        + "  imageSrc?: string;\n"
        + "  translations: Record<SupportedLanguageCode, { language: string; word: string; definition: string }>;\n"
        + "};\n\n"
        + 'export const DEFAULT_LANGUAGE: SupportedLanguageCode = "en";\n\n'
        + "export const WORD_CARDS: WordCard[] = "
        + json.dumps(selected_cards, ensure_ascii=False, indent=2)
        + ";\n"
    )

    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(json.dumps({"cards": len(selected_cards), "output": str(OUTPUT_PATH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
