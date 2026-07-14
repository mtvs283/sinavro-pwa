from __future__ import annotations

import json
from pathlib import Path

import openpyxl


workbook_path = Path("data/translations-24.xlsx")
workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

print(json.dumps({"sheets": workbook.sheetnames}, ensure_ascii=False))

for sheet_name in workbook.sheetnames[:3]:
    sheet = workbook[sheet_name]
    rows: list[list[str]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        rows.append([str(value) if value is not None else "" for value in row[:40]])
        if index >= 8:
            break

    print(json.dumps({"sheet": sheet_name, "rows": rows}, ensure_ascii=False, indent=2))
